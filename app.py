import os
import json
import re
import secrets
import smtplib

from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from werkzeug.security import (
    generate_password_hash,
    check_password_hash
)

from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from io import BytesIO
from functools import wraps

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    session,
    send_file,
    render_template_string
)

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import gspread
from google.oauth2.service_account import Credentials

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.utils import ImageReader

import qrcode


# =========================================================
# FLASK
# =========================================================

app = Flask(__name__)

app.secret_key = os.environ.get(
    "SECRET_KEY",
    "temporary-secret-key-change-me"
)

ADMIN_PASSWORD = os.environ.get(
    "ADMIN_PASSWORD",
    "12345"
)


# =========================================================
# GOOGLE SHEETS
# =========================================================

GOOGLE_SHEET_NAME = "Emek Test 2026"


# =========================================================
# İSTİFADƏÇİLƏR SHEET
# =========================================================

USERS_SHEET_NAME = "İstifadəçilər"

USERS_HEADERS = [
    "İstifadəçi ID",
    "İstifadəçi adı",
    "E-mail",
    "Şifrə hash",
    "Ad və soyad",
    "Qeydiyyat tarixi",
    "Reset kod hash",
    "Reset kod bitmə vaxtı"
]


def get_users_sheet():

    client = get_google_client()

    spreadsheet = client.open(
        GOOGLE_SHEET_NAME
    )

    try:

        sheet = spreadsheet.worksheet(
            USERS_SHEET_NAME
        )

    except gspread.WorksheetNotFound:

        sheet = spreadsheet.add_worksheet(
            title=USERS_SHEET_NAME,
            rows=2000,
            cols=len(USERS_HEADERS)
        )

        sheet.update(
            f"A1:H1",
            [USERS_HEADERS]
        )

    else:

        headers = sheet.row_values(1)

        if not headers:

            sheet.update(
                "A1:H1",
                [USERS_HEADERS]
            )

        else:

            changed = False

            for header in USERS_HEADERS:

                if header not in headers:

                    headers.append(header)

                    changed = True

            if changed:

                sheet.resize(
                    rows=max(
                        sheet.row_count,
                        2000
                    ),
                    cols=max(
                        sheet.col_count,
                        len(headers)
                    )
                )

                sheet.update(
                    "1:1",
                    [headers]
                )

    return sheet


# =========================================================
# İSTİFADƏÇİ ID
# =========================================================

def generate_user_id():

    while True:

        user_id = (
            "USR-"
            +
            secrets.token_hex(5).upper()
        )

        try:

            sheet = get_users_sheet()

            values = sheet.get_all_values()

            existing_ids = set()

            for row in values[1:]:

                if row:

                    existing_ids.add(
                        str(
                            row[0]
                            if len(row) > 0
                            else ""
                        ).strip().upper()
                    )

            if user_id not in existing_ids:

                return user_id

        except Exception as e:

            print(
                "USER ID CHECK ERROR:",
                str(e)
            )

            return user_id


# =========================================================
# İSTİFADƏÇİ MƏLUMATI NORMALİZASİYASI
# =========================================================

def normalize_username(value):

    return str(
        value or ""
    ).strip().lower()


def normalize_email(value):

    return str(
        value or ""
    ).strip().lower()


def normalize_name(value):

    return re.sub(
        r"\s+",
        " ",
        str(value or "").strip()
    )


# =========================================================
# E-MAIL YOXLAMA
# =========================================================

def is_valid_email(email):

    pattern = (
        r"^[A-Za-z0-9._%+-]+@"
        r"[A-Za-z0-9.-]+\.[A-Za-z]{2,}$"
    )

    return bool(
        re.fullmatch(
            pattern,
            email
        )
    )


# =========================================================
# İSTİFADƏÇİ TAP
# =========================================================

def find_user_by_username(username):

    username = normalize_username(
        username
    )

    if not username:
        return None

    try:

        sheet = get_users_sheet()

        values = sheet.get_all_values()

        if len(values) <= 1:
            return None

        for row_number, row in enumerate(
            values[1:],
            start=2
        ):

            if not row:
                continue

            current_username = normalize_username(
                row[1]
                if len(row) > 1
                else ""
            )

            if current_username != username:
                continue

            return {

                "row_number": row_number,

                "user_id": (
                    str(row[0]).strip()
                    if len(row) > 0
                    else ""
                ),

                "username": (
                    str(row[1]).strip()
                    if len(row) > 1
                    else ""
                ),

                "email": (
                    str(row[2]).strip()
                    if len(row) > 2
                    else ""
                ),

                "password_hash": (
                    str(row[3]).strip()
                    if len(row) > 3
                    else ""
                ),

                "name": (
                    str(row[4]).strip()
                    if len(row) > 4
                    else ""
                ),

                "registration_date": (
                    str(row[5]).strip()
                    if len(row) > 5
                    else ""
                ),

                "reset_code_hash": (
                    str(row[6]).strip()
                    if len(row) > 6
                    else ""
                ),

                "reset_code_expiry": (
                    str(row[7]).strip()
                    if len(row) > 7
                    else ""
                )

            }

        return None

    except Exception as e:

        print(
            "FIND USER BY USERNAME ERROR:",
            str(e)
        )

        return None


# =========================================================
# E-MAIL İLƏ İSTİFADƏÇİ TAP
# =========================================================

def find_user_by_email(email):

    email = normalize_email(
        email
    )

    if not email:
        return None

    try:

        sheet = get_users_sheet()

        values = sheet.get_all_values()

        if len(values) <= 1:
            return None

        for row_number, row in enumerate(
            values[1:],
            start=2
        ):

            if not row:
                continue

            current_email = normalize_email(
                row[2]
                if len(row) > 2
                else ""
            )

            if current_email != email:
                continue

            return {

                "row_number": row_number,

                "user_id": (
                    str(row[0]).strip()
                    if len(row) > 0
                    else ""
                ),

                "username": (
                    str(row[1]).strip()
                    if len(row) > 1
                    else ""
                ),

                "email": (
                    str(row[2]).strip()
                    if len(row) > 2
                    else ""
                ),

                "password_hash": (
                    str(row[3]).strip()
                    if len(row) > 3
                    else ""
                ),

                "name": (
                    str(row[4]).strip()
                    if len(row) > 4
                    else ""
                ),

                "registration_date": (
                    str(row[5]).strip()
                    if len(row) > 5
                    else ""
                ),

                "reset_code_hash": (
                    str(row[6]).strip()
                    if len(row) > 6
                    else ""
                ),

                "reset_code_expiry": (
                    str(row[7]).strip()
                    if len(row) > 7
                    else ""
                )

            }

        return None

    except Exception as e:

        print(
            "FIND USER BY EMAIL ERROR:",
            str(e)
        )

        return None


# =========================================================
# İSTİFADƏÇİ YARAT
# =========================================================

def create_user(
    username,
    email,
    password,
    name
):

    username = normalize_username(
        username
    )

    email = normalize_email(
        email
    )

    name = normalize_name(
        name
    )

    if not username:
        return False, "İstifadəçi adı daxil edin."

    if not email:
        return False, "E-mail daxil edin."

    if not name:
        return False, "Ad və soyad daxil edin."

    if not is_valid_email(email):

        return False, (
            "Düzgün e-mail ünvanı daxil edin."
        )

    if len(username) < 3:

        return False, (
            "İstifadəçi adı ən azı 3 simvol olmalıdır."
        )

    if len(password) < 6:

        return False, (
            "Şifrə ən azı 6 simvol olmalıdır."
        )

    if find_user_by_username(username):

        return False, (
            "Bu istifadəçi adı artıq mövcuddur."
        )

    if find_user_by_email(email):

        return False, (
            "Bu e-mail artıq qeydiyyatdan keçib."
        )

    try:

        sheet = get_users_sheet()

        user_id = generate_user_id()

        password_hash = generate_password_hash(
            password
        )

        registration_date = datetime.now(
            ZoneInfo("Asia/Baku")
        ).strftime(
            "%d.%m.%Y %H:%M:%S"
        )

        sheet.append_row(
            [
                user_id,
                username,
                email,
                password_hash,
                name,
                registration_date,
                "",
                ""
            ],
            value_input_option="USER_ENTERED"
        )

        return True, user_id

    except Exception as e:

        print(
            "CREATE USER ERROR:",
            str(e)
        )

        return False, (
            "Qeydiyyat zamanı xəta baş verdi."
        )


# =========================================================
# LOGIN
# =========================================================

def login_user(
    username,
    password
):

    user = find_user_by_username(
        username
    )

    if not user:

        return False, (
            "İstifadəçi adı və ya şifrə yanlışdır."
        )

    password_hash = user.get(
        "password_hash",
        ""
    )

    if not password_hash:

        return False, (
            "İstifadəçi hesabında şifrə məlumatı yoxdur."
        )

    try:

        valid = check_password_hash(
            password_hash,
            password
        )

    except Exception as e:

        print(
            "PASSWORD CHECK ERROR:",
            str(e)
        )

        valid = False

    if not valid:

        return False, (
            "İstifadəçi adı və ya şifrə yanlışdır."
        )

    return True, user


# =========================================================
# E-MAIL SETTINGS
# =========================================================

import os
import resend

resend.api_key = os.environ.get("RESEND_API_KEY")


def send_email(recipient, subject, body):
  try:
    params = {
        "from": "onboarding@resend.dev",
        "to": [recipient],
        "subject": subject,
        "html": body,
    }
    response = resend.Emails.send(params)
    return True
  except Exception as e:
    print("RESEND ERROR:", str(e))
    raise e


# =========================================================
# RESET KODU
# =========================================================

RESET_CODE_MINUTES = 10


def generate_reset_code():

    return (
        f"{secrets.randbelow(1000000):06d}"
    )


def save_reset_code(
    user,
    code
):

    sheet = get_users_sheet()

    expiry = datetime.now(
        ZoneInfo("Asia/Baku")
    ) + timedelta(
        minutes=RESET_CODE_MINUTES
    )

    code_hash = generate_password_hash(
        code
    )

    row_number = user["row_number"]

    sheet.update_cell(
        row_number,
        7,
        code_hash
    )

    sheet.update_cell(
        row_number,
        8,
        expiry.strftime(
            "%d.%m.%Y %H:%M:%S"
        )
    )

    return expiry


def verify_reset_code(
    user,
    code
):

    code = str(
        code or ""
    ).strip()

    if not re.fullmatch(
        r"\d{6}",
        code
    ):

        return False

    code_hash = user.get(
        "reset_code_hash",
        ""
    )

    expiry_text = user.get(
        "reset_code_expiry",
        ""
    )

    if not code_hash or not expiry_text:

        return False

    try:

        if not check_password_hash(
            code_hash,
            code
        ):

            return False

    except Exception as e:

        print(
            "RESET CODE HASH ERROR:",
            str(e)
        )

        return False

    try:
        # Əvvəlcə saniyəli formatı yoxlayırıq, alınmasa saniyəsiz formatı oxuyuruq
        try:
            expiry = datetime.strptime(
                expiry_text,
                "%d.%m.%Y %H:%M:%S"
            ).replace(
                tzinfo=ZoneInfo("Asia/Baku")
            )
        except ValueError:
            expiry = datetime.strptime(
                expiry_text,
                "%d.%m.%Y %H:%M"
            ).replace(
                tzinfo=ZoneInfo("Asia/Baku")
            )

    except Exception as e:

        print(
            "RESET CODE DATE ERROR:",
            str(e)
        )

        return False

    now = datetime.now(
        ZoneInfo("Asia/Baku")
    )

    if now > expiry:

        return False

    return True


def clear_reset_code(user):

    try:

        sheet = get_users_sheet()

        row_number = user["row_number"]

        sheet.update_cell(
            row_number,
            7,
            ""
        )

        sheet.update_cell(
            row_number,
            8,
            ""
        )

    except Exception as e:

        print(
            "CLEAR RESET CODE ERROR:",
            str(e)
        )


# =========================================================
# ACCOUNT PAGE TEMPLATES
# =========================================================

ACCOUNT_BASE_STYLE = """
<style>

    * {
        box-sizing: border-box;
    }

    body {
        margin: 0;
        min-height: 100vh;
        font-family:
            Arial,
            Helvetica,
            sans-serif;
        background:
            linear-gradient(
                135deg,
                #eff6ff,
                #f8fafc
            );
        color: #172033;
        display: flex;
        justify-content: center;
        align-items: center;
        padding: 20px;
    }

    .account-card {
        width: 100%;
        max-width: 500px;
        background: white;
        border-radius: 24px;
        padding: 35px;
        box-shadow:
            0 20px 60px
            rgba(15, 23, 42, .12);
        border: 1px solid #e5e7eb;
    }

    .logo {
        width: 70px;
        height: 70px;
        margin: 0 auto 18px;
        border-radius: 20px;
        background: #123b63;
        color: white;
        display: flex;
        justify-content: center;
        align-items: center;
        font-size: 30px;
        font-weight: 900;
    }

    h1 {
        margin: 0;
        text-align: center;
        color: #123b63;
        font-size: 28px;
    }

    .subtitle {
        text-align: center;
        color: #6b7280;
        margin: 10px 0 25px;
        line-height: 1.5;
    }

    .form-group {
        margin-bottom: 16px;
    }

    label {
        display: block;
        margin-bottom: 7px;
        font-size: 14px;
        font-weight: 700;
        color: #334155;
    }

    input {
        width: 100%;
        padding: 14px 15px;
        border-radius: 12px;
        border: 1px solid #dbe2ea;
        outline: none;
        font-size: 15px;
        background: #f8fafc;
    }

    input:focus {
        border-color: #1e73be;
        background: white;
        box-shadow:
            0 0 0 3px
            rgba(30, 115, 190, .10);
    }

    button {
        width: 100%;
        border: 0;
        border-radius: 12px;
        padding: 14px;
        background: #123b63;
        color: white;
        font-size: 15px;
        font-weight: 800;
        cursor: pointer;
        margin-top: 5px;
    }

    button:hover {
        background: #1e73be;
    }

    .error {
        background: #fee2e2;
        color: #b91c1c;
        border: 1px solid #fecaca;
        border-radius: 12px;
        padding: 12px 14px;
        margin-bottom: 18px;
        line-height: 1.4;
    }

    .success {
        background: #dcfce7;
        color: #166534;
        border: 1px solid #bbf7d0;
        border-radius: 12px;
        padding: 12px 14px;
        margin-bottom: 18px;
        line-height: 1.4;
    }

    .links {
        margin-top: 20px;
        text-align: center;
        line-height: 2;
    }

    .links a {
        color: #1e73be;
        text-decoration: none;
        font-weight: 700;
    }

    .links a:hover {
        text-decoration: underline;
    }

    .code-input {
        text-align: center;
        letter-spacing: 7px;
        font-size: 24px;
        font-weight: 900;
    }

    .note {
        background: #eff6ff;
        border: 1px solid #dbeafe;
        color: #1e3a5f;
        padding: 13px;
        border-radius: 12px;
        margin-bottom: 18px;
        font-size: 13px;
        line-height: 1.5;
    }

    @media(max-width: 600px) {

        .account-card {
            padding: 25px 18px;
            border-radius: 18px;
        }

        h1 {
            font-size: 24px;
        }

    }

</style>
"""


REGISTER_TEMPLATE = """
<!DOCTYPE html>
<html lang="az">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Şəxsi kabinet yarat</title>
    {{ style|safe }}
</head>

<body>

<div class="account-card">

    <div class="logo">ƏM</div>

    <h1>Şəxsi kabinet yarat</h1>

    <div class="subtitle">
        Hesab yaradın və imtahan nəticələrinizi
        şəxsi kabinetinizdən idarə edin.
    </div>

    {% if error %}
        <div class="error">
            {{ error }}
        </div>
    {% endif %}

    {% if success %}
        <div class="success">
            {{ success }}
        </div>
    {% endif %}

    <form method="POST">

        <div class="form-group">

            <label>
                Ad və soyad
            </label>

            <input
                type="text"
                name="name"
                value="{{ name or '' }}"
                placeholder="Ad və soyad"
                required
                autocomplete="name"
            >

        </div>

        <div class="form-group">

            <label>
                İstifadəçi adı
            </label>

            <input
                type="text"
                name="username"
                value="{{ username or '' }}"
                placeholder="Məsələn: perviz"
                required
                autocomplete="username"
            >

        </div>

        <div class="form-group">

            <label>
                E-mail
            </label>

            <input
                type="email"
                name="email"
                value="{{ email or '' }}"
                placeholder="example@gmail.com"
                required
                autocomplete="email"
            >

        </div>

        <div class="form-group">

            <label>
                Şifrə
            </label>

            <input
                type="password"
                name="password"
                placeholder="Ən azı 6 simvol"
                required
                minlength="6"
                autocomplete="new-password"
            >

        </div>

        <div class="form-group">

            <label>
                Şifrəni təkrar daxil edin
            </label>

            <input
                type="password"
                name="password_confirm"
                placeholder="Şifrəni təkrar yazın"
                required
                minlength="6"
                autocomplete="new-password"
            >

        </div>

        <button type="submit">
            ŞƏXSİ KABİNET YARAT
        </button>

    </form>

    <div class="links">

        Artıq hesabınız var?

        <br>

        <a href="{{ url_for('login') }}">
            Şəxsi kabinetə daxil olun
        </a>

        <br>

        <a href="{{ url_for('home') }}">
            ← Test səhifəsinə qayıt
        </a>

    </div>

</div>

</body>
</html>
"""


LOGIN_TEMPLATE = """
<!DOCTYPE html>
<html lang="az">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Şəxsi kabinetə daxil ol</title>
    {{ style|safe }}
</head>

<body>

<div class="account-card">

    <div class="logo">ƏM</div>

    <h1>Şəxsi kabinet</h1>

    <div class="subtitle">
        Hesabınıza daxil olun.
    </div>

    {% if error %}
        <div class="error">
            {{ error }}
        </div>
    {% endif %}

    {% if success %}
        <div class="success">
            {{ success }}
        </div>
    {% endif %}

    <form method="POST">

        <div class="form-group">

            <label>
                İstifadəçi adı
            </label>

            <input
                type="text"
                name="username"
                value="{{ username or '' }}"
                placeholder="İstifadəçi adı"
                required
                autocomplete="username"
            >

        </div>

        <div class="form-group">

            <label>
                Şifrə
            </label>

            <input
                type="password"
                name="password"
                placeholder="Şifrə"
                required
                autocomplete="current-password"
            >

        </div>

        <button type="submit">
            DAXİL OL
        </button>

    </form>

    <div class="links">

        <a href="{{ url_for('forgot_password') }}">
            Şifrəni unutdum
        </a>

        <br>

        Hesabınız yoxdur?

        <a href="{{ url_for('register') }}">
            Şəxsi kabinet yaradın
        </a>

        <br>

        <a href="{{ url_for('home') }}">
            ← Test səhifəsinə qayıt
        </a>

    </div>

</div>

</body>
</html>
"""


FORGOT_TEMPLATE = """
<!DOCTYPE html>
<html lang="az">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Şifrəni unutdum</title>
    {{ style|safe }}
</head>

<body>

<div class="account-card">

    <div class="logo">?</div>

    <h1>Şifrəni unutdum</h1>

    <div class="subtitle">
        Hesabınıza bağlı e-mail ünvanını daxil edin.
    </div>

    {% if error %}
        <div class="error">
            {{ error }}
        </div>
    {% endif %}

    {% if success %}
        <div class="success">
            {{ success }}
        </div>
    {% endif %}

    <form method="POST">

        <div class="form-group">

            <label>
                E-mail
            </label>

            <input
                type="email"
                name="email"
                value="{{ email or '' }}"
                placeholder="example@gmail.com"
                required
                autocomplete="email"
            >

        </div>

        <button type="submit">
            TƏSDİQ KODU GÖNDƏR
        </button>

    </form>

    <div class="links">

        <a href="{{ url_for('login') }}">
            ← Login səhifəsinə qayıt
        </a>

    </div>

</div>

</body>
</html>
"""


RESET_CODE_TEMPLATE = """
<!DOCTYPE html>
<html lang="az">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Təsdiq kodu</title>
    {{ style|safe }}
</head>

<body>

<div class="account-card">

    <div class="logo">✓</div>

    <h1>Təsdiq kodu</h1>

    <div class="subtitle">
        E-mail ünvanınıza göndərilən 6 rəqəmli
        kodu daxil edin.
    </div>

    {% if error %}
        <div class="error">
            {{ error }}
        </div>
    {% endif %}

    {% if success %}
        <div class="success">
            {{ success }}
        </div>
    {% endif %}

    <div class="note">
        Kod {{ minutes }} dəqiqə ərzində etibarlıdır.
    </div>

    <form method="POST">

        <div class="form-group">

            <label>
                6 rəqəmli təsdiq kodu
            </label>

            <input
                class="code-input"
                type="text"
                name="code"
                maxlength="6"
                minlength="6"
                pattern="[0-9]{6}"
                inputmode="numeric"
                autocomplete="one-time-code"
                placeholder="000000"
                required
            >

        </div>

        <button type="submit">
            KODU TƏSDİQLƏ
        </button>

    </form>

    <div class="links">

        <a href="{{ url_for('forgot_password') }}">
            Yeni kod göndər
        </a>

    </div>

</div>

</body>
</html>
"""


RESET_PASSWORD_TEMPLATE = """
<!DOCTYPE html>
<html lang="az">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Yeni şifrə</title>
    {{ style|safe }}
</head>

<body>

<div class="account-card">

    <div class="logo">🔒</div>

    <h1>Yeni şifrə</h1>

    <div class="subtitle">
        Hesabınız üçün yeni şifrə təyin edin.
    </div>

    {% if error %}
        <div class="error">
            {{ error }}
        </div>
    {% endif %}

    <form method="POST">

        <div class="form-group">

            <label>
                Yeni şifrə
            </label>

            <input
                type="password"
                name="password"
                minlength="6"
                placeholder="Ən azı 6 simvol"
                required
                autocomplete="new-password"
            >

        </div>

        <div class="form-group">

            <label>
                Yeni şifrəni təkrar daxil edin
            </label>

            <input
                type="password"
                name="password_confirm"
                minlength="6"
                placeholder="Şifrəni təkrar yazın"
                required
                autocomplete="new-password"
            >

        </div>

        <button type="submit">
            ŞİFRƏNİ YENİLƏ
        </button>

    </form>

</div>

</body>
</html>
"""


# =========================================================
# BÖLMƏLƏR
# =========================================================

SECTIONS = [
    "I Bölmə (Əsas müddəalar)",
    "II Bölmə (Kollektiv müqavilə və sazişin bağlanmasının ümumi qaydaları)",
    "III Bölmə (Əmək müqaviləsinin bağlanması əsasları və qaydası)",
    "IV Bölmə (İş vaxtının növləri və tənzimlənməsi qaydaları)",
    "V Bölmə (İstirahət vaxtı və işçilərin məzuniyyət hüquqları)",
    "VI Bölmə (Əmək normaları və işəmuzd qiymətləri)",
    "VII Bölmə (Əmək və icra intizamının təmin edilməsi qaydaları)",
    "VIII Bölmə (İşəgötürən və işçinin qarşılıqlı maddi məsuliyyətini müəyyən edən hallar)",
    "IX Bölmə (Əməyin mühafizəsi normaları, qaydaları və prinsipləri)",
    "X Bölmə (Qadınların əmək hüququ və onun həyata keçirilməsində təminatları)",
    "XI Bölmə (Kollektiv əmək mübahisələri)",
    "XII Bölmə (İşçilərin sığorta olunmasının tənzimlənməsi)",
    "XIII Bölmə (Əmək məcəlləsinin tələblərinə əməl olunmasına nəzarət. Əmək qanunvericiliyinin pozulmasına görə məsuliyyət)"
]


# =========================================================
# ROMAN RƏQƏMLƏRİ
# =========================================================

ROMAN_NUMBERS = {
    "I": 1,
    "II": 2,
    "III": 3,
    "IV": 4,
    "V": 5,
    "VI": 6,
    "VII": 7,
    "VIII": 8,
    "IX": 9,
    "X": 10,
    "XI": 11,
    "XII": 12,
    "XIII": 13
}


# =========================================================
# BÖLMƏ NORMALİZASİYASI
# =========================================================

def normalize_section(value):

    if value is None:
        return ""

    value = str(value).strip()

    if not value:
        return ""

    value = re.sub(
        r"\s+",
        " ",
        value
    ).strip()

    lower_value = value.lower()

    for section in SECTIONS:

        if lower_value == section.lower():
            return section

    if re.fullmatch(r"\d+", value):

        number = int(value)

        if 1 <= number <= len(SECTIONS):
            return SECTIONS[number - 1]

    match = re.match(
        r"^\s*(\d+)\s*b[öo]lm[əe]\b",
        lower_value
    )

    if match:

        number = int(match.group(1))

        if 1 <= number <= len(SECTIONS):
            return SECTIONS[number - 1]

    first_word = re.split(
        r"[\s\-_()]+",
        value.upper()
    )[0]

    if first_word in ROMAN_NUMBERS:

        number = ROMAN_NUMBERS[first_word]

        if 1 <= number <= len(SECTIONS):
            return SECTIONS[number - 1]

    roman_match = re.match(
        r"^\s*(I{1,3}|IV|V|VI{0,3}|IX|X|XI{0,2}|XII|XIII)"
        r"\s*b[öo]lm[əe]?",
        value,
        re.IGNORECASE
    )

    if roman_match:

        roman = roman_match.group(1).upper()

        if roman in ROMAN_NUMBERS:

            number = ROMAN_NUMBERS[roman]

            return SECTIONS[number - 1]

    roman_start = re.match(
        r"^\s*(I{1,3}|IV|V|VI{0,3}|IX|X|XI{0,2}|XII|XIII)\b",
        value,
        re.IGNORECASE
    )

    if roman_start:

        roman = roman_start.group(1).upper()

        if roman in ROMAN_NUMBERS:

            return SECTIONS[
                ROMAN_NUMBERS[roman] - 1
            ]

    cleaned = (
        lower_value
        .replace("bölmə", "")
        .replace("bölm", "")
        .strip()
    )

    roman_only = cleaned.split("(")[0].strip()

    if roman_only in ROMAN_NUMBERS:

        return SECTIONS[
            ROMAN_NUMBERS[roman_only]
            - 1
        ]

    for section in SECTIONS:

        section_lower = section.lower()

        section_prefix = (
            section_lower
            .split("(")[0]
            .strip()
        )

        if lower_value.startswith(
            section_prefix
        ):
            return section

    return value


# =========================================================
# BÖLMƏ MÜQAYİSƏSİ
# =========================================================

def same_section(value1, value2):

    return (
        normalize_section(value1)
        ==
        normalize_section(value2)
    )


# =========================================================
# GOOGLE CLIENT
# =========================================================

def get_google_client():

    private_key = os.environ.get(
        "GOOGLE_PRIVATE_KEY",
        ""
    )

    if not private_key:

        raise RuntimeError(
            "GOOGLE_PRIVATE_KEY Render Environment Variables "
            "bölməsində yoxdur."
        )

    private_key = private_key.replace(
        "\\n",
        "\n"
    )

    service_account_info = {

        "type": "service_account",

        "project_id": os.environ.get(
            "GOOGLE_PROJECT_ID"
        ),

        "private_key_id": os.environ.get(
            "GOOGLE_PRIVATE_KEY_ID"
        ),

        "private_key": private_key,

        "client_email": os.environ.get(
            "GOOGLE_CLIENT_EMAIL"
        ),

        "token_uri":
            "https://oauth2.googleapis.com/token"
    }

    scopes = [

        "https://www.googleapis.com/auth/spreadsheets",

        "https://www.googleapis.com/auth/drive"

    ]

    credentials = Credentials.from_service_account_info(
        service_account_info,
        scopes=scopes
    )

    return gspread.authorize(credentials)


# =========================================================
# NƏTİCƏLƏR SHEET
# =========================================================

def get_sheet():

    client = get_google_client()

    spreadsheet = client.open(
        GOOGLE_SHEET_NAME
    )

    try:

        sheet = spreadsheet.worksheet(
            "Nəticələr"
        )

    except gspread.WorksheetNotFound:

        sheet = spreadsheet.add_worksheet(
            title="Nəticələr",
            rows=1000,
            cols=13
        )

    required_headers = [

        "№",
        "Ad və soyad",
        "Bölmə",
        "Düzgün cavab",
        "Ümumi sual",
        "Səhv cavab",
        "Nəticə",
        "Tarix",
        "Status",
        "Müddət",
        "Başlama vaxtı",
        "Bitmə vaxtı",
        "Sertifikat №"

    ]

    headers = sheet.row_values(1)

    if not headers:

        sheet.update(
            "A1:M1",
            [required_headers]
        )

    else:

        changed = False

        for header in required_headers:

            if header not in headers:

                headers.append(header)

                changed = True

        if changed:

            sheet.resize(
                rows=max(
                    sheet.row_count,
                    1000
                ),
                cols=max(
                    sheet.col_count,
                    len(headers)
                )
            )

            sheet.update(
                "1:1",
                [headers]
            )

    return sheet


# =========================================================
# SERTİFİKATLAR SHEET
# =========================================================

CERTIFICATES_SHEET_NAME = "Sertifikatlar"

CERTIFICATES_HEADERS = [

    "Sertifikat №",
    "Ad və soyad",
    "Bölmə",
    "Düzgün cavab",
    "Ümumi sual",
    "Nəticə",
    "Tarix",
    "Müddət",
    "Status",
    "Başlama vaxtı",
    "Bitmə vaxtı"

]


def get_certificates_sheet():

    client = get_google_client()

    spreadsheet = client.open(
        GOOGLE_SHEET_NAME
    )

    try:

        sheet = spreadsheet.worksheet(
            CERTIFICATES_SHEET_NAME
        )

    except gspread.WorksheetNotFound:

        sheet = spreadsheet.add_worksheet(
            title=CERTIFICATES_SHEET_NAME,
            rows=2000,
            cols=len(CERTIFICATES_HEADERS)
        )

        sheet.update(
            "A1:K1",
            [CERTIFICATES_HEADERS]
        )

    else:

        headers = sheet.row_values(1)

        if not headers:

            sheet.update(
                "A1:K1",
                [CERTIFICATES_HEADERS]
            )

    return sheet


# =========================================================
# UNİKAL SERTİFİKAT NÖMRƏSİ
# =========================================================

def generate_certificate_number():

    characters = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"

    while True:

        random_part = "".join(
            secrets.choice(characters)
            for _ in range(6)
        )

        certificate_number = (
            f"ƏMƏK-{datetime.now().year}-{random_part}"
        )

        try:

            sheet = get_certificates_sheet()

            values = sheet.get_all_values()

            existing_numbers = {

                str(row[0]).strip().upper()

                for row in values[1:]
                if row

            }

            if (
                certificate_number.upper()
                not in existing_numbers
            ):

                return certificate_number

        except Exception as e:

            print(
                "CERTIFICATE NUMBER CHECK ERROR:",
                str(e)
            )

            return certificate_number


# =========================================================
# SERTİFİKAT SAXLA
# =========================================================

def save_certificate(
    certificate_number,
    name,
    section,
    correct,
    total,
    percent,
    created_at,
    duration_text,
    status,
    start_time_text,
    end_time_text
):

    sheet = get_certificates_sheet()

    row_data = [

        certificate_number,
        name,
        section,
        correct,
        total,
        f"{percent}%",
        created_at,
        duration_text,
        "Etibarlı",
        start_time_text,
        end_time_text

    ]

    sheet.append_row(
        row_data,
        value_input_option="USER_ENTERED"
    )

    return True


# =========================================================
# SERTİFİKAT TAP
# =========================================================

def find_certificate(certificate_number):

    certificate_number = str(
        certificate_number or ""
    ).strip().upper()

    if not certificate_number:
        return None

    try:

        sheet = get_certificates_sheet()

        values = sheet.get_all_values()

        if len(values) <= 1:
            return None

        for row in values[1:]:

            if not row:
                continue

            current_number = str(
                row[0]
                if len(row) > 0
                else ""
            ).strip().upper()

            if current_number != certificate_number:
                continue

            return {

                "certificate_number": current_number,

                "name": (
                    str(row[1]).strip()
                    if len(row) > 1
                    else ""
                ),

                "section": (
                    str(row[2]).strip()
                    if len(row) > 2
                    else ""
                ),

                "correct": (
                    str(row[3]).strip()
                    if len(row) > 3
                    else "0"
                ),

                "total": (
                    str(row[4]).strip()
                    if len(row) > 4
                    else "0"
                ),

                "percent": (
                    str(row[5]).strip()
                    if len(row) > 5
                    else "0%"
                ),

                "date": (
                    str(row[6]).strip()
                    if len(row) > 6
                    else ""
                ),

                "duration": (
                    str(row[7]).strip()
                    if len(row) > 7
                    else ""
                ),

                "status": (
                    str(row[8]).strip()
                    if len(row) > 8
                    else "Etibarlı"
                ),

                "start_time": (
                    str(row[9]).strip()
                    if len(row) > 9
                    else "-"
                ),

                "end_time": (
                    str(row[10]).strip()
                    if len(row) > 10
                    else "-"
                )

            }

        return None

    except Exception as e:

        print(
            "FIND CERTIFICATE ERROR:",
            str(e)
        )

        return None


# =========================================================
# SUALLAR SHEET
# =========================================================

QUESTIONS_SHEET_NAME = "Suallar"

QUESTIONS_HEADERS = [
    "Bölmə",
    "question",
    "a",
    "b",
    "c",
    "d",
    "answer"
]


def get_questions_sheet():

    client = get_google_client()

    spreadsheet = client.open(
        GOOGLE_SHEET_NAME
    )

    try:

        sheet = spreadsheet.worksheet(
            QUESTIONS_SHEET_NAME
        )

    except gspread.WorksheetNotFound:

        sheet = spreadsheet.add_worksheet(
            title=QUESTIONS_SHEET_NAME,
            rows=5000,
            cols=7
        )

        sheet.update(
            "A1:G1",
            [QUESTIONS_HEADERS]
        )

    else:

        headers = sheet.row_values(1)

        if not headers:

            sheet.update(
                "A1:G1",
                [QUESTIONS_HEADERS]
            )

    return sheet


# =========================================================
# BÜTÜN SUALLARI YÜKLƏ
# =========================================================

def load_all_questions():

    try:

        sheet = get_questions_sheet()

        values = sheet.get_all_values()

        if len(values) <= 1:
            return []

        questions = []

        for row_number, row in enumerate(
            values[1:],
            start=2
        ):

            if not row:
                continue

            if len(row) < 7:
                continue

            raw_section = str(
                row[0]
            ).strip()

            section = normalize_section(
                raw_section
            )

            if section not in SECTIONS:

                print(
                    f"SUAL {row_number}: "
                    f"Naməlum bölmə -> {raw_section}"
                )

                continue

            question_text = str(
                row[1]
            ).strip()

            if not question_text:
                continue

            answer = str(
                row[6]
            ).strip().upper()

            if len(answer) > 1:
                answer = answer[0]

            if answer not in [
                "A",
                "B",
                "C",
                "D"
            ]:

                print(
                    f"SUAL {row_number}: "
                    f"Yanlış cavab -> {answer}"
                )

                continue

            questions.append({

                "section": section,

                "question": question_text,

                "a": str(row[2]).strip(),

                "b": str(row[3]).strip(),

                "c": str(row[4]).strip(),

                "d": str(row[5]).strip(),

                "answer": answer

            })

        print(
            "SUALLAR YÜKLƏNDİ:",
            len(questions)
        )

        return questions

    except Exception as e:

        print(
            "LOAD ALL QUESTIONS ERROR:",
            str(e)
        )

        return []


# =========================================================
# SEÇİLMİŞ BÖLMƏNİN SUALLARI
# =========================================================

def load_questions(section=None):

    all_questions = load_all_questions()

    if section is None:
        return all_questions

    normalized_requested = normalize_section(
        section
    )

    if normalized_requested not in SECTIONS:
        return []

    result = [

        q
        for q in all_questions

        if normalize_section(
            q.get("section", "")
        )
        ==
        normalized_requested

    ]

    print(
        "BÖLMƏ:",
        normalized_requested,
        "| SUAL SAYI:",
        len(result)
    )

    return result


# =========================================================
# KODLAR SHEET
# =========================================================

CODES_SHEET_NAME = "Kodlar"

CODES_HEADERS = [
    "Kod",
    "Bölmə",
    "Status",
    "Ad və soyad",
    "İstifadə vaxtı"
]


def get_codes_sheet():

    client = get_google_client()

    spreadsheet = client.open(
        GOOGLE_SHEET_NAME
    )

    try:

        sheet = spreadsheet.worksheet(
            CODES_SHEET_NAME
        )

    except gspread.WorksheetNotFound:

        sheet = spreadsheet.add_worksheet(
            title=CODES_SHEET_NAME,
            rows=2000,
            cols=5
        )

        sheet.update(
            "A1:E1",
            [CODES_HEADERS]
        )

    else:

        headers = sheet.row_values(1)

        if not headers:

            sheet.update(
                "A1:E1",
                [CODES_HEADERS]
            )

    return sheet


# =========================================================
# KODU İSTİFADƏ ET
# =========================================================

def use_access_code(
    code,
    selected_section,
    name
):

    code = str(
        code or ""
    ).strip()

    selected_section = normalize_section(
        selected_section
    )

    name = str(
        name or ""
    ).strip()

    if not code:

        return False, (
            "Giriş kodunu daxil edin."
        )

    if selected_section not in SECTIONS:

        return False, (
            "Zəhmət olmasa bölmə seçin."
        )

    try:

        sheet = get_codes_sheet()

        values = sheet.get_all_values()

        if len(values) <= 1:

            return False, (
                "Hazırda aktiv giriş kodu yoxdur."
            )

        for row_number, row in enumerate(
            values[1:],
            start=2
        ):

            if not row:
                continue

            sheet_code = str(
                row[0] if len(row) > 0 else ""
            ).strip()

            if (
                sheet_code.upper()
                !=
                code.upper()
            ):
                continue

            sheet_section = normalize_section(
                row[1] if len(row) > 1 else ""
            )

            status = str(
                row[2] if len(row) > 2 else ""
            ).strip().lower()

            if status != "aktiv":

                return False, (
                    "Bu giriş kodu artıq istifadə olunub."
                )

            if not same_section(
                sheet_section,
                selected_section
            ):

                return False, (
                    "Bu giriş kodu seçdiyiniz bölmə üçün "
                    "nəzərdə tutulmayıb."
                )

            current_time = datetime.now(
                ZoneInfo("Asia/Baku")
            ).strftime(
                "%d.%m.%Y %H:%M:%S"
            )

            sheet.update_cell(
                row_number,
                3,
                "İstifadə olunub"
            )

            sheet.update_cell(
                row_number,
                4,
                name
            )

            sheet.update_cell(
                row_number,
                5,
                current_time
            )

            return True, ""

        return False, (
            "Giriş kodu yanlışdır."
        )

    except Exception as e:

        print(
            "ACCESS CODE ERROR:",
            str(e)
        )

        return False, (
            "Giriş kodu yoxlanılarkən xəta baş verdi."
        )


# =========================================================
# ADMIN AUTH
# =========================================================

def admin_required(function):

    @wraps(function)
    def wrapper(*args, **kwargs):

        if not session.get("admin"):

            return redirect(
                url_for("admin_login")
            )

        return function(
            *args,
            **kwargs
        )

    return wrapper


# =========================================================
# ACCOUNT AUTH
# =========================================================

def account_required(function):

    @wraps(function)
    def wrapper(*args, **kwargs):

        if not session.get("user_id"):

            return redirect(
                url_for("login")
            )

        return function(
            *args,
            **kwargs
        )

    return wrapper


# =========================================================
# REGISTER
# =========================================================

@app.route(
    "/register",
    methods=["GET", "POST"]
)
def register():

    if session.get("user_id"):

        return redirect(
            url_for("cabinet")
        )

    if request.method == "POST":

        name = normalize_name(
            request.form.get(
                "name",
                ""
            )
        )

        username = normalize_username(
            request.form.get(
                "username",
                ""
            )
        )

        email = normalize_email(
            request.form.get(
                "email",
                ""
            )
        )

        password = request.form.get(
            "password",
            ""
        )

        password_confirm = request.form.get(
            "password_confirm",
            ""
        )

        if password != password_confirm:

            return render_template_string(
                REGISTER_TEMPLATE,
                style=ACCOUNT_BASE_STYLE,
                error="Şifrələr eyni deyil.",
                name=name,
                username=username,
                email=email
            )

        success, result = create_user(
            username=username,
            email=email,
            password=password,
            name=name
        )

        if not success:

            return render_template_string(
                REGISTER_TEMPLATE,
                style=ACCOUNT_BASE_STYLE,
                error=result,
                name=name,
                username=username,
                email=email
            )

        return redirect(
            url_for(
                "login",
                registered="1"
            )
        )

    return render_template_string(
        REGISTER_TEMPLATE,
        style=ACCOUNT_BASE_STYLE
    )

# =========================================================
# LOGIN
# =========================================================

@app.route(
    "/login",
    methods=["GET", "POST"]
)
def login():

    if session.get("user_id"):

        return redirect(
            url_for("cabinet")
        )

    registered = request.args.get(
        "registered"
    )

    success = None

    if registered == "1":

        success = (
            "Hesabınız uğurla yaradıldı. "
            "İndi şəxsi kabinetə daxil ola bilərsiniz."
        )

    if request.method == "POST":

        username = normalize_username(
            request.form.get(
                "username",
                ""
            )
        )

        password = request.form.get(
            "password",
            ""
        )

        valid, result = login_user(
            username,
            password
        )

        if not valid:

            return render_template_string(
                LOGIN_TEMPLATE,
                style=ACCOUNT_BASE_STYLE,
                error=result,
                success=success,
                username=username
            )

        session["user_id"] = result["user_id"]

        session["username"] = result["username"]

        session["email"] = result["email"]

        session["account_name"] = result["name"]

        session.modified = True

        return redirect(
            url_for("cabinet")
        )

    return render_template_string(
        LOGIN_TEMPLATE,
        style=ACCOUNT_BASE_STYLE,
        success=success
    )


# =========================================================
# LOGOUT
# =========================================================

@app.route(
    "/logout"
)
def logout():

    session.clear()

    return redirect(
        url_for("home")
    )


# =========================================================
# FORGOT PASSWORD
# =========================================================

@app.route(
    "/forgot-password",
    methods=["GET", "POST"]
)
def forgot_password():

    if request.method == "POST":

        email = normalize_email(
            request.form.get(
                "email",
                ""
            )
        )

        if not is_valid_email(email):

            return render_template_string(
                FORGOT_TEMPLATE,
                style=ACCOUNT_BASE_STYLE,
                error=(
                    "Düzgün e-mail ünvanı daxil edin."
                ),
                email=email
            )

        user = find_user_by_email(
            email
        )

        if not user:

            return render_template_string(
                FORGOT_TEMPLATE,
                style=ACCOUNT_BASE_STYLE,
                error=(
                    "Bu e-mail ilə qeydiyyatdan keçmiş "
                    "istifadəçi tapılmadı."
                ),
                email=email
            )

        code = generate_reset_code()

        try:

            save_reset_code(
                user,
                code
            )

            email_body = f"""
Salam, {user["name"]}!

Şifrəni yeniləmək üçün təsdiq kodunuz:

{code}

Bu kod {RESET_CODE_MINUTES} dəqiqə ərzində etibarlıdır.

Əgər bu sorğunu siz göndərməmisinizsə,
bu e-maili nəzərə almayın.

Əmək Məcəlləsi test sistemi
"""

            send_email(
                recipient=user["email"],
                subject="Şifrə yeniləmə təsdiq kodu",
                body=email_body
            )

            session["reset_user_id"] = user["user_id"]

            session["reset_email"] = user["email"]

            session.modified = True

            return redirect(
                url_for("verify_reset")
            )

        except Exception as e:

            print(
                "SEND RESET EMAIL ERROR:",
                str(e)
            )

            return render_template_string(
                FORGOT_TEMPLATE,
                style=ACCOUNT_BASE_STYLE,
                error=(
                    "Təsdiq kodu göndərilərkən xəta baş verdi. "
                    "Resend API açarını yoxlayın."
                ),
                email=email
            )

    return render_template_string(
        FORGOT_TEMPLATE,
        style=ACCOUNT_BASE_STYLE
    )


# =========================================================
# VERIFY RESET CODE
# =========================================================

@app.route(
    "/verify-reset",
    methods=["GET", "POST"]
)
def verify_reset():

    user_id = session.get(
        "reset_user_id"
    )

    if not user_id:

        return redirect(
            url_for("forgot_password")
        )

    user = find_user_by_username(
        session.get(
            "reset_username",
            ""
        )
    )

    if not user:

        try:

            sheet = get_users_sheet()

            values = sheet.get_all_values()

            for row_number, row in enumerate(
                values[1:],
                start=2
            ):

                if not row:
                    continue

                current_id = str(
                    row[0]
                    if len(row) > 0
                    else ""
                ).strip()

                if current_id != str(user_id).strip():
                    continue

                user = {

                    "row_number": row_number,

                    "user_id": current_id,

                    "username": (
                        str(row[1]).strip()
                        if len(row) > 1
                        else ""
                    ),

                    "email": (
                        str(row[2]).strip()
                        if len(row) > 2
                        else ""
                    ),

                    "password_hash": (
                        str(row[3]).strip()
                        if len(row) > 3
                        else ""
                    ),

                    "name": (
                        str(row[4]).strip()
                        if len(row) > 4
                        else ""
                    ),

                    "registration_date": (
                        str(row[5]).strip()
                        if len(row) > 5
                        else ""
                    ),

                    "reset_code_hash": (
                        str(row[6]).strip()
                        if len(row) > 6
                        else ""
                    ),

                    "reset_code_expiry": (
                        str(row[7]).strip()
                        if len(row) > 7
                        else ""
                    )

                }

                break

        except Exception as e:

            print(
                "VERIFY RESET USER ERROR:",
                str(e)
            )

    if not user:

        session.pop(
            "reset_user_id",
            None
        )

        session.pop(
            "reset_email",
            None
        )

        return redirect(
            url_for("forgot_password")
        )

    if request.method == "POST":

        code = str(
            request.form.get(
                "code",
                ""
            )
        ).strip()

        if not verify_reset_code(
            user,
            code
        ):

            return render_template_string(
                RESET_CODE_TEMPLATE,
                style=ACCOUNT_BASE_STYLE,
                error=(
                    "Kod yanlışdır və ya müddəti bitib."
                ),
                minutes=RESET_CODE_MINUTES
            )

        session["reset_verified"] = True

        session.modified = True

        return redirect(
            url_for("reset_password")
        )

    return render_template_string(
        RESET_CODE_TEMPLATE,
        style=ACCOUNT_BASE_STYLE,
        minutes=RESET_CODE_MINUTES
    )


# =========================================================
# RESET PASSWORD
# =========================================================

@app.route(
    "/reset-password",
    methods=["GET", "POST"]
)
def reset_password():

    if not session.get(
        "reset_verified"
    ):

        return redirect(
            url_for("forgot_password")
        )

    user_id = session.get(
        "reset_user_id"
    )

    if not user_id:

        return redirect(
            url_for("forgot_password")
        )

    user = None

    try:

        sheet = get_users_sheet()

        values = sheet.get_all_values()

        for row_number, row in enumerate(
            values[1:],
            start=2
        ):

            if not row:
                continue

            current_id = str(
                row[0]
                if len(row) > 0
                else ""
            ).strip()

            if current_id != str(
                user_id
            ).strip():
                continue

            user = {

                "row_number": row_number,

                "user_id": current_id,

                "username": (
                    str(row[1]).strip()
                    if len(row) > 1
                    else ""
                ),

                "email": (
                    str(row[2]).strip()
                    if len(row) > 2
                    else ""
                ),

                "password_hash": (
                    str(row[3]).strip()
                    if len(row) > 3
                    else ""
                ),

                "name": (
                    str(row[4]).strip()
                    if len(row) > 4
                    else ""
                )

            }

            break

    except Exception as e:

        print(
            "RESET PASSWORD USER ERROR:",
            str(e)
        )

    if not user:

        session.clear()

        return redirect(
            url_for("forgot_password")
        )

    if request.method == "POST":

        password = request.form.get(
            "password",
            ""
        )

        password_confirm = request.form.get(
            "password_confirm",
            ""
        )

        if len(password) < 6:

            return render_template_string(
                RESET_PASSWORD_TEMPLATE,
                style=ACCOUNT_BASE_STYLE,
                error=(
                    "Yeni şifrə ən azı 6 simvol olmalıdır."
                )
            )

        if password != password_confirm:

            return render_template_string(
                RESET_PASSWORD_TEMPLATE,
                style=ACCOUNT_BASE_STYLE,
                error="Şifrələr eyni deyil."
            )

        password_hash = generate_password_hash(
            password
        )

        try:

            sheet = get_users_sheet()

            sheet.update_cell(
                user["row_number"],
                4,
                password_hash
            )

            sheet.update_cell(
                user["row_number"],
                7,
                ""
            )

            sheet.update_cell(
                user["row_number"],
                8,
                ""
            )

            session.clear()

            return redirect(
                url_for(
                    "login",
                    registered="0"
                )
            )

        except Exception as e:

            print(
                "RESET PASSWORD SAVE ERROR:",
                str(e)
            )

            return render_template_string(
                RESET_PASSWORD_TEMPLATE,
                style=ACCOUNT_BASE_STYLE,
                error=(
                    "Yeni şifrə yadda saxlanılarkən "
                    "xəta baş verdi."
                )
            )

    return render_template_string(
        RESET_PASSWORD_TEMPLATE,
        style=ACCOUNT_BASE_STYLE
    )


# =========================================================
# HOME
# =========================================================

@app.route("/", methods=["GET", "POST"])
def home():

    if "user_id" not in session:
        return redirect(url_for("login"))

    # Burada user_name yox, account_name istifadə etmək daha düzgündür
    logged_in_name = session.get(
        "account_name",
        ""
    ).strip()

    if request.method == "POST":

        name = request.form.get(
            "name",
            ""
        ).strip()

        if not name:
            name = logged_in_name

        access_code = request.form.get(
            "access_code",
            ""
        ).strip()

        selected_section = normalize_section(
            request.form.get(
                "section",
                ""
            ).strip()
        )

        # =========================================
        # AD VƏ SOYAD YOXLAMASI
        # =========================================

        if not name:

            return render_template(
                "home.html",
                error="Ad və soyad daxil edin.",
                name=name,
                access_code=access_code,
                selected_section=selected_section,
                sections=SECTIONS,
                user_name=logged_in_name
            )

        # =========================================
        # BÖLMƏ YOXLAMASI
        # =========================================

        if selected_section not in SECTIONS:

            return render_template(
                "home.html",
                error="Bölmə seçin.",
                name=name,
                access_code=access_code,
                selected_section=selected_section,
                sections=SECTIONS,
                user_name=logged_in_name
            )

        # =========================================
        # GİRİŞ KODU YOXLAMASI
        # =========================================

        if not access_code:

            return render_template(
                "home.html",
                error="Giriş kodunu daxil edin.",
                name=name,
                access_code=access_code,
                selected_section=selected_section,
                sections=SECTIONS,
                user_name=logged_in_name
            )

        # =========================================
        # SUALLARI YOXLAYIRIQ
        # =========================================

        section_questions = load_questions(
            selected_section
        )

        if not section_questions:

            return render_template(
                "home.html",
                error=f"{selected_section} üzrə hazırda sual mövcud deyil.",
                name=name,
                access_code=access_code,
                selected_section=selected_section,
                sections=SECTIONS,
                user_name=logged_in_name
            )

        # =========================================
        # GİRİŞ KODUNU YOXLAYIRIQ
        # =========================================

        code_valid, error_message = use_access_code(
            access_code,
            selected_section,
            name
        )

        if not code_valid:

            return render_template(
                "home.html",
                error=error_message,
                name=name,
                access_code=access_code,
                selected_section=selected_section,
                sections=SECTIONS,
                user_name=logged_in_name
            )

        # =========================================
        # HESAB MƏLUMATLARINI QORU
        # =========================================

        account_user_id = session.get(
            "user_id"
        )

        account_username = session.get(
            "username"
        )

        account_email = session.get(
            "email"
        )

        account_name = session.get(
            "account_name"
        )

        # =========================================
        # İMTAHAN SESSION-UNU TƏMİZLƏ
        # =========================================

        session.clear()

        # =========================================
        # HESAB MƏLUMATLARINI GERİ QAYTAR
        # =========================================

        if account_user_id:

            session["user_id"] = account_user_id
            session["username"] = account_username
            session["email"] = account_email
            session["account_name"] = account_name

        # =========================================
        # İMTAHAN MƏLUMATLARI
        # =========================================

        session["name"] = name
        session["access_code"] = access_code
        session["section"] = selected_section
        session["answers"] = {}

        session["exam_finished"] = False

        start_time = datetime.now(
            ZoneInfo("Asia/Baku")
        )

        session["exam_start_time"] = (
            start_time.timestamp()
        )

        session.modified = True

        # =========================================
        # İMTAHANA KEÇ
        # =========================================

        return redirect(
            url_for(
                "question",
                n=0
            )
        )

    # =========================================
    # NORMAL SƏHİFƏ AÇILIŞI
    # =========================================

    return render_template(
        "home.html",
        sections=SECTIONS,
        name=logged_in_name,
        selected_section="",
        user_name=logged_in_name
    )

# =========================================================
# QUESTION
# =========================================================

@app.route(
    "/question/<int:n>",
    methods=["GET", "POST"]
)
def question(n):

    if "name" not in session:
        return redirect(url_for("home"))

    if "access_code" not in session:
        return redirect(url_for("home"))

    if "section" not in session:
        return redirect(url_for("home"))

    if session.get("exam_finished"):
        return redirect(url_for("finish"))

    selected_section = normalize_section(
        session.get("section")
    )

    current_questions = load_questions(
        selected_section
    )

    total = len(current_questions)

    if total == 0:

        return render_template(
            "home.html",
            error=(
                f"{selected_section} üzrə "
                f"hazırda sual yoxdur."
            ),
            sections=SECTIONS
        )

    if n < 0:

        return redirect(
            url_for(
                "question",
                n=0
            )
        )

    if n >= total:

        return redirect(
            url_for("finish")
        )

    q = current_questions[n]

    if request.method == "POST":

        selected = request.form.get(
            "answer"
        )

        answers = session.get(
            "answers",
            {}
        )

        if selected in [
            "A",
            "B",
            "C",
            "D"
        ]:

            answers[str(n)] = selected

        else:

            answers.pop(
                str(n),
                None
            )

        session["answers"] = answers

        session.modified = True

        if n + 1 >= total:

            return redirect(
                url_for("finish")
            )

        return redirect(
            url_for(
                "question",
                n=n + 1
            )
        )

    exam_start_time = session.get(
        "exam_start_time"
    )

    return render_template(
        "question.html",
        q=q,
        n=n,
        total=total,
        name=session["name"],
        section=selected_section,
        exam_start_time=exam_start_time
    )


# =========================================================
# FINISH
# =========================================================

@app.route(
    "/finish",
    methods=["GET", "POST"]
)
def finish():

    if "name" not in session:
        return redirect(url_for("home"))

    if "access_code" not in session:
        return redirect(url_for("home"))

    if "section" not in session:
        return redirect(url_for("home"))

    if session.get("exam_finished"):

        return render_template(
            "finish.html",
            name=session.get("name"),
            section=session.get("section"),
            correct=session.get("finish_correct", 0),
            total=session.get("finish_total", 0),
            percent=session.get("finish_percent", 0),
            answered=session.get("finish_answered", 0),
            wrong=session.get("finish_wrong", 0),
            unanswered=session.get("finish_unanswered", 0),
            status=session.get("finish_status", ""),
            duration_text=session.get(
                "finish_duration_text",
                ""
            ),
            duration_minutes=session.get(
                "finish_duration_minutes",
                0
            ),
            duration_seconds=session.get(
                "finish_duration_seconds",
                0
            ),
            start_time_text=session.get(
                "finish_start_time_text",
                "-"
            ),
            end_time_text=session.get(
                "finish_end_time_text",
                "-"
            ),
            certificate_number=session.get(
                "certificate_number",
                ""
            ),
            question_results=session.get(
                "finish_question_results",
                []
            )
        )

    answers = session.get(
        "answers",
        {}
    )

    selected_section = normalize_section(
        session.get("section")
    )

    current_questions = load_questions(
        selected_section
    )

    total = len(current_questions)

    if total == 0:
        return redirect(url_for("home"))

    correct = 0

    answered = len(answers)

    exam_start_timestamp = session.get(
        "exam_start_time"
    )

    now_datetime = datetime.now(
        ZoneInfo("Asia/Baku")
    )

    now_timestamp = now_datetime.timestamp()

    if exam_start_timestamp is not None:

        exam_start_timestamp = float(
            exam_start_timestamp
        )

        exam_start_datetime = datetime.fromtimestamp(
            exam_start_timestamp,
            ZoneInfo("Asia/Baku")
        )

        start_time_text = (
            exam_start_datetime.strftime(
                "%d.%m.%Y %H:%M:%S"
            )
        )

        elapsed_seconds = int(
            now_timestamp
            -
            exam_start_timestamp
        )

        if elapsed_seconds < 0:
            elapsed_seconds = 0

    else:

        start_time_text = "-"
        elapsed_seconds = 0

    end_time_text = (
        now_datetime.strftime(
            "%d.%m.%Y %H:%M:%S"
        )
    )

    duration_minutes = (
        elapsed_seconds // 60
    )

    duration_seconds = (
        elapsed_seconds % 60
    )

    duration_text = (
        f"{duration_minutes} dəq "
        f"{duration_seconds} san"
    )

    question_results = []

    for index, q in enumerate(
        current_questions
    ):

        selected = answers.get(
            str(index)
        )

        correct_letter = q["answer"]

        correct_text = q[
            correct_letter.lower()
        ]

        if selected in [
            "A",
            "B",
            "C",
            "D"
        ]:

            selected_text = q[
                selected.lower()
            ]

        else:

            selected_text = (
                "Cavab verilməyib"
            )

        if selected == correct_letter:

            is_correct = True
            is_wrong = False
            is_empty = False

            correct += 1

        elif selected is None:

            is_correct = False
            is_wrong = False
            is_empty = True

        else:

            is_correct = False
            is_wrong = True
            is_empty = False

        question_results.append({

            "number": index + 1,

            "question": q["question"],

            "selected": selected,

            "selected_text": selected_text,

            "correct_answer": correct_letter,

            "correct_text": correct_text,

            "is_correct": is_correct,

            "is_wrong": is_wrong,

            "is_empty": is_empty

        })

    wrong = answered - correct

    unanswered = total - answered

    percent = (
        round(
            correct / total * 100
        )
        if total
        else 0
    )

    name = session["name"]

    created_at = now_datetime.strftime(
        "%d.%m.%Y %H:%M:%S"
    )

    if answered == total:

        status = "Tamamlandı"

    else:

        status = (
            f"Yarımçıq bitirildi "
            f"({answered}/{total})"
        )

    # =====================================================
    # SERTİFİKAT NÖMRƏSİ
    # =====================================================

    certificate_number = generate_certificate_number()

    # =====================================================
    # NƏTİCƏLƏR SHEET
    # =====================================================

    try:

    sheet = get_sheet()

    all_values = sheet.get_all_values()

    number = len(all_values)

    user_id = str(
        session.get(
            "user_id",
            ""
        )
    ).strip()

    row_data = [

        number,
        user_id,
        name,
        selected_section,
        correct,
        total,
        wrong,
        f"{percent}%",
        created_at,
        status,
        duration_text,
        start_time_text,
        end_time_text,
        certificate_number

    ]

    sheet.append_row(
        row_data,
        value_input_option="USER_ENTERED"
    )

    print(
        "GOOGLE SHEETS: nəticə əlavə edildi."
    )

except Exception as e:

    print(
        "GOOGLE SHEETS ERROR:",
        str(e)
    )

    # =====================================================
    # SERTİFİKAT SHEET
    # =====================================================

    try:

        save_certificate(

            certificate_number=certificate_number,

            name=name,

            section=selected_section,

            correct=correct,

            total=total,

            percent=percent,

            created_at=created_at,

            duration_text=duration_text,

            status=status,

            start_time_text=start_time_text,

            end_time_text=end_time_text

        )

        print(
            "SERTİFİKAT YARADILDI:",
            certificate_number
        )

    except Exception as e:

        print(
            "CERTIFICATE SAVE ERROR:",
            str(e)
        )

    session["exam_finished"] = True

    session["certificate_number"] = certificate_number

    session["finish_correct"] = correct
    session["finish_total"] = total
    session["finish_percent"] = percent
    session["finish_answered"] = answered
    session["finish_wrong"] = wrong
    session["finish_unanswered"] = unanswered
    session["finish_status"] = status
    session["finish_duration_text"] = duration_text
    session["finish_duration_minutes"] = duration_minutes
    session["finish_duration_seconds"] = duration_seconds
    session["finish_start_time_text"] = start_time_text
    session["finish_end_time_text"] = end_time_text
    session["finish_question_results"] = question_results

    session.modified = True

    return render_template(
        "finish.html",
        name=name,
        section=selected_section,
        correct=correct,
        total=total,
        percent=percent,
        answered=answered,
        wrong=wrong,
        unanswered=unanswered,
        status=status,
        duration_text=duration_text,
        duration_minutes=duration_minutes,
        duration_seconds=duration_seconds,
        start_time_text=start_time_text,
        end_time_text=end_time_text,
        certificate_number=certificate_number,
        question_results=question_results
    )


# =========================================================
# PDF FONT
# =========================================================

def register_pdf_font():

    possible_fonts = [

        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",

        "/usr/share/fonts/dejavu/DejaVuSans.ttf",

        "C:/Windows/Fonts/DejaVuSans.ttf",

        "C:/Windows/Fonts/arial.ttf",

        "/Library/Fonts/Arial.ttf"

    ]

    for font_path in possible_fonts:

        if os.path.exists(font_path):

            try:

                pdfmetrics.registerFont(
                    TTFont(
                        "CertificateFont",
                        font_path
                    )
                )

                return "CertificateFont"

            except Exception as e:

                print(
                    "PDF FONT ERROR:",
                    str(e)
                )

    return "Helvetica"


# =========================================================
# PDF MƏTNİ SƏTİRLƏRƏ BÖL
# =========================================================

def draw_wrapped_centered(
    pdf,
    text,
    font_name,
    font_size,
    center_x,
    start_y,
    max_width,
    line_gap=18
):

    words = str(text).split()

    lines = []
    current_line = ""

    for word in words:

        test_line = (
            f"{current_line} {word}"
            if current_line
            else word
        )

        if pdf.stringWidth(
            test_line,
            font_name,
            font_size
        ) <= max_width:

            current_line = test_line

        else:

            if current_line:
                lines.append(current_line)

            current_line = word

    if current_line:
        lines.append(current_line)

    y = start_y

    for line in lines:

        pdf.drawCentredString(
            center_x,
            y,
            line
        )

        y -= line_gap

    return y


# =========================================================
# QR KOD YARAT
# =========================================================

def create_qr_image(data):

    qr = qrcode.QRCode(

        version=None,

        error_correction=qrcode.constants.ERROR_CORRECT_H,

        box_size=8,

        border=2

    )

    qr.add_data(data)

    qr.make(
        fit=True
    )

    image = qr.make_image(
        fill_color="black",
        back_color="white"
    )

    output = BytesIO()

    image.save(
        output,
        format="PNG"
    )

    output.seek(0)

    return output


# =========================================================
# PDF SERTİFİKAT YARATMA
# =========================================================

def build_certificate_pdf(
    name,
    section,
    correct,
    total,
    percent,
    status,
    duration_text,
    end_time_text,
    certificate_number
):

    verify_url = (
        request.host_url.rstrip("/")
        +
        url_for(
            "verify_certificate",
            certificate_number=certificate_number
        )
    )

    qr_output = create_qr_image(
        verify_url
    )

    qr_image = ImageReader(
        qr_output
    )

    output = BytesIO()

    page_width, page_height = landscape(A4)

    pdf = canvas.Canvas(
        output,
        pagesize=landscape(A4)
    )

    font_name = register_pdf_font()

    pdf.setTitle(
        "Əmək Məcəlləsi - Test Sertifikatı"
    )

    dark_blue = colors.HexColor(
        "#123B63"
    )

    blue = colors.HexColor(
        "#1E73BE"
    )

    light_blue = colors.HexColor(
        "#EAF4FF"
    )

    gold = colors.HexColor(
        "#D4A017"
    )

    green = colors.HexColor(
        "#159957"
    )

    dark_gray = colors.HexColor(
        "#444444"
    )

    white = colors.white

    pdf.setFillColor(
        colors.HexColor("#F7FAFD")
    )

    pdf.rect(
        0,
        0,
        page_width,
        page_height,
        fill=1,
        stroke=0
    )

    pdf.setFillColor(
        dark_blue
    )

    pdf.roundRect(
        24,
        24,
        page_width - 48,
        page_height - 48,
        18,
        fill=0,
        stroke=1
    )

    pdf.setLineWidth(4)

    pdf.setStrokeColor(
        gold
    )

    pdf.roundRect(
        34,
        34,
        page_width - 68,
        page_height - 68,
        14,
        fill=0,
        stroke=1
    )

    pdf.setLineWidth(1)

    pdf.setStrokeColor(
        blue
    )

    pdf.roundRect(
        46,
        46,
        page_width - 92,
        page_height - 92,
        10,
        fill=0,
        stroke=1
    )

    pdf.setFillColor(
        dark_blue
    )

    pdf.roundRect(
        47,
        page_height - 105,
        page_width - 94,
        48,
        8,
        fill=1,
        stroke=0
    )

    pdf.setFillColor(
        white
    )

    pdf.setFont(
        font_name,
        22
    )

    pdf.drawCentredString(
        page_width / 2,
        page_height - 87,
        "ƏMƏK MƏCƏLLƏSİ SINAQ PLATFORMASI"
    )

    pdf.setFillColor(
        dark_blue
    )

    pdf.setFont(
        font_name,
        30
    )

    pdf.drawCentredString(
        page_width / 2,
        page_height - 145,
        "S E R T İ F İ K A T"
    )

    pdf.setStrokeColor(
        gold
    )

    pdf.setLineWidth(3)

    pdf.line(
        page_width / 2 - 105,
        page_height - 158,
        page_width / 2 + 105,
        page_height - 158
    )

    badge_x = 75
    badge_y = page_height - 190

    pdf.setFillColor(
        gold
    )

    pdf.circle(
        badge_x,
        badge_y,
        27,
        fill=1,
        stroke=0
    )

    pdf.setFillColor(
        white
    )

    pdf.circle(
        badge_x,
        badge_y,
        20,
        fill=0,
        stroke=1
    )

    pdf.setFont(
        font_name,
        16
    )

    pdf.drawCentredString(
        badge_x,
        badge_y - 6,
        "✓"
    )

    pdf.setFillColor(
        dark_gray
    )

    pdf.setFont(
        font_name,
        13
    )

    pdf.drawCentredString(
        page_width / 2,
        page_height - 195,
        "Bu sertifikat aşağıdakı iştirakçıya təqdim olunur:"
    )

    pdf.setFillColor(
        dark_blue
    )

    pdf.setFont(
        font_name,
        25
    )

    safe_display_name = str(
        name or ""
    ).strip()

    if len(safe_display_name) > 45:

        safe_display_name = (
            safe_display_name[:45]
            + "..."
        )

    pdf.drawCentredString(
        page_width / 2,
        page_height - 235,
        safe_display_name
    )

    name_width = pdf.stringWidth(
        safe_display_name,
        font_name,
        25
    )

    pdf.setStrokeColor(
        gold
    )

    pdf.setLineWidth(2)

    pdf.line(
        max(
            80,
            (page_width - name_width) / 2
        ),
        page_height - 245,
        min(
            page_width - 80,
            (page_width + name_width) / 2
        ),
        page_height - 245
    )

    pdf.setFillColor(
        dark_gray
    )

    pdf.setFont(
        font_name,
        12
    )

    pdf.drawCentredString(
        page_width / 2,
        page_height - 275,
        "İştirak etdiyi bölmə:"
    )

    pdf.setFillColor(
        blue
    )

    pdf.setFont(
        font_name,
        13
    )

    section_bottom_y = draw_wrapped_centered(
        pdf,
        section,
        font_name,
        13,
        page_width / 2,
        page_height - 297,
        page_width - 250,
        17
    )

    result_box_y = section_bottom_y - 82

    pdf.setFillColor(
        light_blue
    )

    pdf.setStrokeColor(
        blue
    )

    pdf.setLineWidth(1.5)

    pdf.roundRect(
        100,
        result_box_y,
        page_width - 200,
        75,
        12,
        fill=1,
        stroke=1
    )

    pdf.setFillColor(
        green
    )

    pdf.setFont(
        font_name,
        27
    )

    pdf.drawString(
        130,
        result_box_y + 42,
        f"{percent}%"
    )

    pdf.setFillColor(
        dark_gray
    )

    pdf.setFont(
        font_name,
        11
    )

    pdf.drawString(
        130,
        result_box_y + 22,
        "Yekun nəticə"
    )

    pdf.setFillColor(
        dark_blue
    )

    pdf.setFont(
        font_name,
        16
    )

    pdf.drawCentredString(
        page_width / 2,
        result_box_y + 43,
        f"{correct} / {total}"
    )

    pdf.setFont(
        font_name,
        10
    )

    pdf.setFillColor(
        dark_gray
    )

    pdf.drawCentredString(
        page_width / 2,
        result_box_y + 22,
        "Düzgün cavab"
    )

    pdf.setFillColor(
        dark_blue
    )

    pdf.setFont(
        font_name,
        10
    )

    pdf.drawRightString(
        page_width - 130,
        result_box_y + 43,
        status
    )

    pdf.setFillColor(
        dark_gray
    )

    pdf.drawRightString(
        page_width - 130,
        result_box_y + 22,
        f"Müddət: {duration_text}"
    )

    pdf.setStrokeColor(
        colors.HexColor("#D8E2EC")
    )

    pdf.setLineWidth(1)

    pdf.line(
        100,
        92,
        page_width - 100,
        92
    )

    pdf.setFillColor(
        dark_gray
    )

    pdf.setFont(
        font_name,
        10
    )

    pdf.drawString(
        105,
        70,
        "Testin bitmə tarixi:"
    )

    pdf.setFillColor(
        dark_blue
    )

    pdf.drawString(
        205,
        70,
        end_time_text
    )

    pdf.setFillColor(
        dark_gray
    )

    pdf.setFont(
        font_name,
        8
    )

    pdf.drawString(
        105,
        52,
        "Sertifikat №:"
    )

    pdf.setFillColor(
        dark_blue
    )

    pdf.setFont(
        font_name,
        9
    )

    pdf.drawString(
        170,
        52,
        certificate_number
    )

    qr_size = 58

    qr_x = page_width - 125

    qr_y = 105

    pdf.drawImage(
        qr_image,
        qr_x,
        qr_y,
        width=qr_size,
        height=qr_size,
        preserveAspectRatio=True,
        mask="auto"
    )

    pdf.setFillColor(
        dark_gray
    )

    pdf.setFont(
        font_name,
        7
    )

    pdf.drawCentredString(
        qr_x + qr_size / 2,
        qr_y - 9,
        "Sertifikatı yoxla"
    )

    pdf.setFillColor(
        dark_gray
    )

    pdf.setFont(
        font_name,
        8
    )

    pdf.drawCentredString(
        page_width / 2,
        52,
        "Əmək Məcəlləsi üzrə elektron test sistemi"
    )

    pdf.setFillColor(
        blue
    )

    pdf.setFont(
        font_name,
        7
    )

    pdf.drawCentredString(
        page_width / 2,
        40,
        "Elektron sertifikat"
    )

    pdf.showPage()

    pdf.save()

    output.seek(0)

    return output


# =========================================================
# SERTİFİKAT YÜKLƏ
# =========================================================

@app.route(
    "/download-certificate"
)
def download_certificate():

    if "name" not in session:
        return redirect(url_for("home"))

    if "section" not in session:
        return redirect(url_for("home"))

    if not session.get("exam_finished"):
        return redirect(url_for("finish"))

    certificate_number = session.get(
        "certificate_number",
        ""
    )

    if not certificate_number:
        return redirect(url_for("finish"))

    certificate = find_certificate(
        certificate_number
    )

    if not certificate:
        return redirect(url_for("finish"))

    output = build_certificate_pdf(

        name=certificate["name"],

        section=certificate["section"],

        correct=certificate["correct"],

        total=certificate["total"],

        percent=certificate["percent"].replace(
            "%",
            ""
        ),

        status=certificate["status"],

        duration_text=certificate["duration"],

        end_time_text=certificate["end_time"],

        certificate_number=certificate[
            "certificate_number"
        ]

    )

    safe_name = "".join(
        c
        for c in certificate["name"]
        if c.isalnum()
        or c in (
            " ",
            "_",
            "-"
        )
    ).strip()

    if not safe_name:
        safe_name = "istifadeci"

    filename = (
        f"{safe_name}_sertifikat.pdf"
    )

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf"
    )


# =========================================================
# KABİNETDƏN SERTİFİKAT YÜKLƏ
# =========================================================

@app.route(
    "/download-certificate/<certificate_number>"
)
def download_certificate_by_number(
    certificate_number
):

    if not session.get("user_id"):

        return redirect(
            url_for("login")
        )

    certificate = find_certificate(
        certificate_number
    )

    if not certificate:
        return redirect(
            url_for("cabinet")
        )

    session_name = str(
        session.get(
            "account_name",
            session.get("name", "")
        )
    ).strip().lower()

    certificate_name = str(
        certificate.get("name", "")
    ).strip().lower()

    if session_name != certificate_name:

        return redirect(
            url_for("cabinet")
        )

    output = build_certificate_pdf(

        name=certificate["name"],

        section=certificate["section"],

        correct=certificate["correct"],

        total=certificate["total"],

        percent=certificate["percent"].replace(
            "%",
            ""
        ),

        status=certificate["status"],

        duration_text=certificate["duration"],

        end_time_text=certificate["end_time"],

        certificate_number=certificate[
            "certificate_number"
        ]

    )

    safe_name = "".join(
        c
        for c in certificate["name"]
        if c.isalnum()
        or c in (
            " ",
            "_",
            "-"
        )
    ).strip()

    if not safe_name:
        safe_name = "istifadeci"

    filename = (
        f"{safe_name}_sertifikat.pdf"
    )

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf"
    )


# =========================================================
# ŞƏXSİ KABİNET
# =========================================================

@app.route("/cabinet")
@account_required
def cabinet():

    account_name = str(
        session.get(
            "account_name",
            ""
        )
    ).strip()

    username = str(
        session.get(
            "username",
            ""
        )
    ).strip()

    email = str(
        session.get(
            "email",
            ""
        )
    ).strip()

    user_id = str(
        session.get(
            "user_id",
            ""
        )
    ).strip()

    results = []

    try:

        sheet = get_sheet()

        values = sheet.get_all_records()

        for row in values:

            row_user_id = str(
                row.get(
                    "İstifadəçi ID",
                    ""
                )
            ).strip()

            if row_user_id == user_id:

                results.append(row)

    except Exception as e:

        print(
            "CABINET RESULTS ERROR:",
            str(e)
        )

    # Ən son nəticə əvvəl görünsün
    results.reverse()

    latest_result = None

    if results:
        latest_result = results[0]

    return render_template(
        "cabinet.html",
        name=account_name,
        username=username,
        email=email,
        results=results,
        certificates=[],
        latest_result=latest_result,
        phase_one=True
    )


# =========================================================
# SERTİFİKAT YOXLAMA
# =========================================================

@app.route(
    "/verify/<certificate_number>"
)
def verify_certificate(
    certificate_number
):

    certificate = find_certificate(
        certificate_number
    )

    if not certificate:

        return render_template_string(
            VERIFY_TEMPLATE,
            valid=False,
            certificate_number=certificate_number
        )

    return render_template_string(
        VERIFY_TEMPLATE,
        valid=True,
        certificate=certificate
    )


# =========================================================
# VERİFİKASİYA SƏHİFƏSİ
# =========================================================

VERIFY_TEMPLATE = """
<!DOCTYPE html>
<html lang="az">

<head>

    <meta charset="UTF-8">

    <meta name="viewport"
          content="width=device-width, initial-scale=1.0">

    <title>Sertifikat yoxlanışı</title>

    <style>

        * {
            box-sizing: border-box;
        }

        body {
            margin: 0;
            min-height: 100vh;
            background:
                linear-gradient(
                    135deg,
                    #eff6ff,
                    #f8fafc
                );
            font-family:
                Arial,
                Helvetica,
                sans-serif;
            color: #172033;
            display: flex;
            justify-content: center;
            align-items: center;
            padding: 20px;
        }

        .card {
            width: 100%;
            max-width: 650px;
            background: white;
            border-radius: 24px;
            padding: 35px;
            box-shadow:
                0 20px 60px
                rgba(15, 23, 42, .12);
            border: 1px solid #e5e7eb;
        }

        .icon {
            width: 78px;
            height: 78px;
            border-radius: 50%;
            margin: 0 auto 18px;
            display: flex;
            justify-content: center;
            align-items: center;
            font-size: 38px;
            font-weight: 900;
        }

        .valid-icon {
            background: #dcfce7;
            color: #15803d;
        }

        .invalid-icon {
            background: #fee2e2;
            color: #b91c1c;
        }

        h1 {
            text-align: center;
            margin: 0;
            font-size: 28px;
            color: #123b63;
        }

        .subtitle {
            text-align: center;
            color: #6b7280;
            margin: 10px 0 28px;
        }

        .certificate-number {
            background: #eff6ff;
            color: #173b73;
            border-radius: 12px;
            padding: 13px;
            text-align: center;
            font-family: monospace;
            font-weight: 800;
            letter-spacing: 1px;
            margin-bottom: 25px;
            word-break: break-word;
        }

        .info {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 12px;
        }

        .item {
            background: #f8fafc;
            border: 1px solid #e5e7eb;
            border-radius: 12px;
            padding: 14px;
        }

        .label {
            color: #6b7280;
            font-size: 12px;
            margin-bottom: 5px;
        }

        .value {
            color: #123b63;
            font-weight: 800;
            font-size: 14px;
            word-break: break-word;
        }

        .footer {
            margin-top: 25px;
            text-align: center;
            color: #9ca3af;
            font-size: 12px;
        }

        @media(max-width: 600px) {

            .card {
                padding: 25px 18px;
                border-radius: 18px;
            }

            .info {
                grid-template-columns: 1fr;
            }

            h1 {
                font-size: 24px;
            }

        }

    </style>

</head>

<body>

    <div class="card">

        {% if valid %}

            <div class="icon valid-icon">
                ✓
            </div>

            <h1>
                Sertifikat etibarlıdır
            </h1>

            <div class="subtitle">
                Bu sertifikat sistemdə təsdiqlənmişdir.
            </div>

            <div class="certificate-number">
                {{ certificate.certificate_number }}
            </div>

            <div class="info">

                <div class="item">

                    <div class="label">
                        Ad və soyad
                    </div>

                    <div class="value">
                        {{ certificate.name }}
                    </div>

                </div>

                <div class="item">

                    <div class="label">
                        Bölmə
                    </div>

                    <div class="value">
                        {{ certificate.section }}
                    </div>

                </div>

                <div class="item">

                    <div class="label">
                        Nəticə
                    </div>

                    <div class="value">
                        {{ certificate.percent }}
                    </div>

                </div>

                <div class="item">

                    <div class="label">
                        Düzgün cavab sayı
                    </div>

                    <div class="value">
                        {{ certificate.correct }}
                        /
                        {{ certificate.total }}
                    </div>

                </div>

                <div class="item">

                    <div class="label">
                        Test tarixi
                    </div>

                    <div class="value">
                        {{ certificate.date }}
                    </div>

                </div>

                <div class="item">

                    <div class="label">
                        Müddət
                    </div>

                    <div class="value">
                        {{ certificate.duration }}
                    </div>

                </div>

                <div class="item">

                    <div class="label">
                        Status
                    </div>

                    <div class="value">
                        {{ certificate.status }}
                    </div>

                </div>

                <div class="item">

                    <div class="label">
                        Bitmə vaxtı
                    </div>

                    <div class="value">
                        {{ certificate.end_time }}
                    </div>

                </div>

            </div>

        {% else %}

            <div class="icon invalid-icon">
                ✕
            </div>

            <h1>
                Sertifikat etibarsızdır
            </h1>

            <div class="subtitle">
                Sertifikat tapılmadı.
            </div>

            <div class="certificate-number">
                {{ certificate_number }}
            </div>

        {% endif %}

        <div class="footer">
            Əmək Məcəlləsi üzrə elektron test sistemi
        </div>

    </div>

</body>

</html>
"""


# =========================================================
# ADMIN LOGIN
# =========================================================

@app.route(
    "/admin/login",
    methods=["GET", "POST"]
)
def admin_login():

    if request.method == "POST":

        password = request.form.get(
            "password",
            ""
        )

        if password == ADMIN_PASSWORD:

            session["admin"] = True

            return redirect(
                url_for("admin")
            )

        return render_template(
            "admin_login.html",
            error="Parol yanlışdır."
        )

    return render_template(
        "admin_login.html"
    )


# =========================================================
# ADMIN LOGOUT
# =========================================================

@app.route("/cabinet/logout")
def cabinet_logout():

    session.clear()

    return redirect(
        url_for("home")
    )


@app.route("/admin/logout")
def admin_logout():

    session.pop(
        "admin",
        None
    )

    return redirect(
        url_for("admin_login")
    )


# =========================================================
# ADMIN PANEL
# =========================================================

@app.route(
    "/admin"
)
@admin_required
def admin():

    try:

        sheet = get_sheet()

        values = sheet.get_all_records()

    except Exception as e:

        print(
            "ADMIN GOOGLE SHEETS ERROR:",
            str(e)
        )

        values = []

    questions = load_all_questions()

    try:

        codes_sheet = get_codes_sheet()

        codes = codes_sheet.get_all_records()

    except Exception as e:

        print(
            "ADMIN CODES ERROR:",
            str(e)
        )

        codes = []

    try:

        certificates_sheet = get_certificates_sheet()

        certificates = (
            certificates_sheet.get_all_records()
        )

    except Exception as e:

        print(
            "ADMIN CERTIFICATES ERROR:",
            str(e)
        )

        certificates = []

    try:

        users_sheet = get_users_sheet()

        users = (
            users_sheet.get_all_records()
        )

    except Exception as e:

        print(
            "ADMIN USERS ERROR:",
            str(e)
        )

        users = []

    return render_template(
        "admin.html",
        results=values,
        questions=questions,
        sections=SECTIONS,
        codes=codes,
        certificates=certificates,
        users=users
    )


# =========================================================
# CREATE CODE
# =========================================================

def create_code():

    code = request.form.get(
        "code",
        ""
    ).strip()

    section = normalize_section(
        request.form.get(
            "section",
            ""
        ).strip()
    )

    if not code:
        return redirect(url_for("admin"))

    if section not in SECTIONS:
        return redirect(url_for("admin"))

    try:

        sheet = get_codes_sheet()

        values = sheet.get_all_values()

        for row in values[1:]:

            if not row:
                continue

            existing_code = str(
                row[0]
                if len(row) > 0
                else ""
            ).strip()

            if (
                existing_code
                and
                existing_code.upper()
                ==
                code.upper()
            ):

                return redirect(
                    url_for("admin")
                )

        sheet.append_row(
            [
                code,
                section,
                "Aktiv",
                "",
                ""
            ],
            value_input_option="USER_ENTERED"
        )

    except Exception as e:

        print(
            "CREATE CODE ERROR:",
            str(e)
        )

    return redirect(
        url_for("admin")
    )


# =========================================================
# ADMIN ADD CODE
# =========================================================

@app.route(
    "/admin/add-code",
    methods=["POST"]
)
@admin_required
def admin_add_code():

    return create_code()


# =========================================================
# ADMIN CREATE CODE
# =========================================================

@app.route(
    "/admin/create-code",
    methods=["POST"]
)
@admin_required
def admin_create_code():

    return create_code()


# =========================================================
# IMPORT QUESTIONS
# =========================================================

@app.route(
    "/admin/import-questions",
    methods=["POST"]
)
@admin_required
def import_questions():

    file = request.files.get(
        "questions_file"
    )

    if not file:
        return redirect(url_for("admin"))

    filename = (
        file.filename or ""
    ).lower()

    new_questions = []

    try:

        if (
            filename.endswith(".xlsx")
            or filename.endswith(".xls")
        ):

            workbook = load_workbook(
                file,
                data_only=True
            )

            sheet = workbook.active

            for row in sheet.iter_rows(
                min_row=2,
                values_only=True
            ):

                if not row:
                    continue

                raw_section = (
                    str(row[0]).strip()
                    if len(row) > 0
                    and row[0] is not None
                    else ""
                )

                section = normalize_section(
                    raw_section
                )

                if section not in SECTIONS:
                    continue

                question_text = (
                    str(row[1]).strip()
                    if len(row) > 1
                    and row[1] is not None
                    else ""
                )

                if not question_text:
                    continue

                option_a = (
                    str(row[2]).strip()
                    if len(row) > 2
                    and row[2] is not None
                    else ""
                )

                option_b = (
                    str(row[3]).strip()
                    if len(row) > 3
                    and row[3] is not None
                    else ""
                )

                option_c = (
                    str(row[4]).strip()
                    if len(row) > 4
                    and row[4] is not None
                    else ""
                )

                option_d = (
                    str(row[5]).strip()
                    if len(row) > 5
                    and row[5] is not None
                    else ""
                )

                answer = (
                    str(row[6]).strip().upper()
                    if len(row) > 6
                    and row[6] is not None
                    else ""
                )

                if len(answer) > 1:
                    answer = answer[0]

                if answer not in [
                    "A",
                    "B",
                    "C",
                    "D"
                ]:
                    continue

                new_questions.append({

                    "section": section,

                    "question": question_text,

                    "a": option_a,

                    "b": option_b,

                    "c": option_c,

                    "d": option_d,

                    "answer": answer

                })

        elif filename.endswith(".json"):

            loaded_questions = json.load(
                file
            )

            if isinstance(
                loaded_questions,
                list
            ):

                for q in loaded_questions:

                    if not isinstance(
                        q,
                        dict
                    ):
                        continue

                    section = normalize_section(
                        q.get(
                            "section",
                            ""
                        )
                    )

                    if section not in SECTIONS:
                        continue

                    question_text = str(
                        q.get(
                            "question",
                            ""
                        )
                    ).strip()

                    if not question_text:
                        continue

                    answer = str(
                        q.get(
                            "answer",
                            ""
                        )
                    ).strip().upper()

                    if len(answer) > 1:
                        answer = answer[0]

                    if answer not in [
                        "A",
                        "B",
                        "C",
                        "D"
                    ]:
                        continue

                    new_questions.append({

                        "section": section,

                        "question": question_text,

                        "a": str(
                            q.get(
                                "a",
                                ""
                            )
                        ).strip(),

                        "b": str(
                            q.get(
                                "b",
                                ""
                            )
                        ).strip(),

                        "c": str(
                            q.get(
                                "c",
                                ""
                            )
                        ).strip(),

                        "d": str(
                            q.get(
                                "d",
                                ""
                            )
                        ).strip(),

                        "answer": answer

                    })

        else:

            print(
                "IMPORT ERROR: "
                "Fayl formatı dəstəklənmir."
            )

            return redirect(
                url_for("admin")
            )

        if not new_questions:

            return redirect(
                url_for("admin")
            )

        questions_sheet = get_questions_sheet()

        questions_sheet.clear()

        questions_sheet.update(
            "A1:G1",
            [QUESTIONS_HEADERS]
        )

        rows = []

        for q in new_questions:

            rows.append([

                q["section"],
                q["question"],
                q["a"],
                q["b"],
                q["c"],
                q["d"],
                q["answer"]

            ])

        questions_sheet.update(
            f"A2:G{len(rows) + 1}",
            rows
        )

        print(
            "IMPORT UĞURLU:",
            len(new_questions),
            "sual saxlanıldı."
        )

    except Exception as e:

        print(
            "IMPORT ERROR:",
            str(e)
        )

    return redirect(
        url_for("admin")
    )


# =========================================================
# EXCEL EXPORT
# =========================================================

@app.route(
    "/admin/export"
)
@admin_required
def admin_export():

    try:

        sheet = get_sheet()

        results = sheet.get_all_records()

    except Exception as e:

        print(
            "EXPORT GOOGLE SHEETS ERROR:",
            str(e)
        )

        results = []

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Test nəticələri"

    headers = [
    "№",
    "İstifadəçi ID",
    "Ad və soyad",
    "Bölmə",
    "Düzgün",
    "Ümumi",
    "Səhv",
    "Faiz",
    "Tarix",
    "Status",
    "Müddət",
    "Başlama vaxtı",
    "Bitmə vaxtı",
    "Sertifikat nömrəsi"
]

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="123B63"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for column, header in enumerate(
        headers,
        1
    ):

        cell = worksheet.cell(
            row=1,
            column=column,
            value=header
        )

        cell.font = header_font

        cell.fill = header_fill

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = thin_border

    for row_number, result in enumerate(
        results,
        2
    ):

        values = [

            result.get(
                "№",
                row_number - 1
            ),

            result.get(
                "Ad və soyad",
                ""
            ),

            result.get(
                "Bölmə",
                ""
            ),

            result.get(
                "Düzgün cavab",
                0
            ),

            result.get(
                "Ümumi sual",
                0
            ),

            result.get(
                "Səhv cavab",
                0
            ),

            result.get(
                "Nəticə",
                ""
            ),

            result.get(
                "Tarix",
                ""
            ),

            result.get(
                "Status",
                ""
            ),

            result.get(
                "Müddət",
                ""
            ),

            result.get(
                "Başlama vaxtı",
                ""
            ),

            result.get(
                "Bitmə vaxtı",
                ""
            ),

            result.get(
                "Sertifikat №",
                ""
            )

        ]

        for column, value in enumerate(
            values,
            1
        ):

            cell = worksheet.cell(
                row=row_number,
                column=column,
                value=value
            )

            cell.border = thin_border

            cell.alignment = Alignment(
                vertical="center"
            )

    widths = {

        "A": 8,
        "B": 30,
        "C": 55,
        "D": 18,
        "E": 18,
        "F": 15,
        "G": 15,
        "H": 25,
        "I": 30,
        "J": 18,
        "K": 25,
        "L": 25,
        "M": 22

    }

    for column, width in widths.items():

        worksheet.column_dimensions[
            column
        ].width = width

    worksheet.freeze_panes = "A2"

    output = BytesIO()

    workbook.save(
        output
    )

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=(
            "emek_mecellesi_test_neticeleri.xlsx"
        ),
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )


# =========================================================
# HEALTH CHECK
# =========================================================

@app.route(
    "/health"
)
def health():

    return "OK", 200


# =========================================================
# RUN
# =========================================================

if __name__ == "__main__":

    app.run(
        host="0.0.0.0",
        port=int(
            os.environ.get(
                "PORT",
                5000
            )
        )
    )
